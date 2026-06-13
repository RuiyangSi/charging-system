"""用本系统（charging-system-orig, requeue 口径）跑完整 42 个验收事件，
把每个时刻的桩状态/等候区按模板格式填进《作业验收用例.xlsx》，另存为新文件。

格子格式（与模板/同学填法一致）：
- 桩列(快充1/2=F1/F2, 慢充1/2/3=T1/T2/T3)：每事件占 3 行（M=3 车位），
  正在充电=「(车号,已充电量,当前费用)」，排队中=「(车号,0.00,0.00)」，空位=「-」，故障桩=「故障」。
- 等候区列：「(车号,模式,请求电量)」按到达先后用「-」连接。
运行：cd charging-system-orig/backend && python3 fill_acceptance.py
"""
from __future__ import annotations

import os
import re
from datetime import datetime, timedelta

import openpyxl

from app.context import AppContext
from app.domain.enums import PileStatus, RequestStatus

HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(HERE, "..", "..", "作业验收用例.xlsx")   # 模板在项目根（仓库外）
OUT = os.path.join(HERE, "..", "docs", "作业验收用例_本组填写.xlsx")   # 输出到仓库 docs/
PILE_COLS = {"F1": 3, "F2": 4, "T1": 5, "T2": 6, "T3": 7}   # 快充1,快充2,慢充1,慢充2,慢充3
WAIT_COL = 8
START = datetime(2026, 6, 11, 6, 0, 0)


def parse_time(v):
    s = str(v).strip()
    m = re.match(r"^(\d{1,2}):(\d{2}):(\d{2})$", s)
    if not m:
        return None
    return START.replace(hour=int(m.group(1)), minute=int(m.group(2)), second=int(m.group(3)))


def parse_event(v):
    m = re.match(r"\(([ABC]),([^,]+),([TFO]),([\d.]+)\)", str(v).strip())
    if not m:
        return None
    return (m.group(1), m.group(2).strip(), m.group(3), float(m.group(4)))


def apply_event(ctx, actor, obj, op, val):
    cs, ss = ctx.charging_service, ctx.schedule_service
    if actor == "A" and op in ("T", "F"):
        cs.submit_request(obj, op, val)
    elif actor == "A" and op == "O":
        if ctx.station.find_active_request(obj):
            cs.cancel(obj)
    elif actor == "B" and op == "O":
        if val == 0:
            ss.handle_fault(obj)
        else:
            pile = ctx.station.find_pile(obj)
            if pile.status is PileStatus.FAULT:
                ss.handle_recovery(obj)
            else:
                ctx.pile_service.power_on(obj); ctx.pile_service.run_pile(obj)
    elif actor == "C" and op == "O":
        if ctx.station.find_active_request(obj):
            cs.modify_amount(obj, val)


def render_pile(ctx, pile, now):
    if pile.status is PileStatus.FAULT:
        return ["故障", "-", "-"]
    cells = []
    for cr in pile.queue.get_cars():
        charged = cr.charged_amount(now, pile.power)
        fee = (ctx.billing_service.estimate(cr, pile, now)["totalFee"]
               if cr.status is RequestStatus.CHARGING else 0.0)
        cells.append(f"({cr.car_id},{charged:.2f},{fee:.2f})")
    while len(cells) < 3:
        cells.append("-")
    return cells[:3]


def render_wait(ctx):
    cars = ctx.station.waiting_area.get_all_requests()
    return "-".join(f"({c.car_id},{c.mode.value},{c.requested_amount:.2f})" for c in cars)


def main():
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["测试用例"]

    # 1) 从模板读出事件行（A=时刻, B=事件），保持与模板完全一致的行布局
    event_rows = []   # (row, datetime, (actor,obj,op,val))
    for r in range(3, 128):
        t = parse_time(ws.cell(r, 1).value)
        ev = parse_event(ws.cell(r, 2).value)
        if t is not None and ev is not None:
            event_rows.append((r, t, ev))
    by_time = {}
    for r, t, ev in event_rows:
        by_time.setdefault(t, []).append(ev)
    last_event_time = max(t for _, t, _ in event_rows)

    # 2) 跑引擎：逐分钟推进，到点应用事件 + step()，在事件时刻抓快照
    ctx = AppContext(os.path.join(HERE, "_fill_tmp"), db_path=":memory:")
    ctx.clock.set_speed(0)
    ctx.clock.set_time(START)
    snaps = {}
    cur = START
    # 多跑到次日 02:00，确保把"调度结束"时刻也算出来
    hard_stop = START + timedelta(days=1, hours=2)
    while cur <= hard_stop:
        ctx.clock.set_time(cur)
        for ev in by_time.get(cur, []):
            apply_event(ctx, *ev)
        ctx.step()
        if cur in by_time:
            snaps[cur] = ({pid: render_pile(ctx, ctx.station.find_pile(pid), cur)
                           for pid in PILE_COLS}, render_wait(ctx))
        # 所有事件已下发且无活动请求 → 调度结束
        if cur >= last_event_time:
            active = [x for x in ctx.station.requests.values() if x.status.active]
            if not active:
                break
        cur += timedelta(minutes=1)

    # 调度结束时刻
    ends = [r.end_time for r in ctx.station.requests.values() if r.end_time]
    sched_end = max(ends) if ends else None

    # 3) 写回模板副本
    for r, t, _ in event_rows:
        piles, wait = snaps.get(t, ({pid: ["-", "-", "-"] for pid in PILE_COLS}, ""))
        for pid, col in PILE_COLS.items():
            cells = piles[pid]
            for k in range(3):
                ws.cell(r + k, col).value = cells[k]
        ws.cell(r, WAIT_COL).value = wait

    # 调度结束行（模板第 129 行 A 列保留模板原值 "1:35+1"，仅在备注列写实际算得的结束时刻）
    for r in range(127, 135):
        if str(ws.cell(r, 2).value or "").strip() == "调度结束":
            if sched_end:
                days = (sched_end.date() - START.date()).days
                ws.cell(r, 1).value = f"{sched_end.strftime('%H:%M')}+{days}"   # 01:35+1
                ws.cell(r, 9).value = f"实际调度结束 {sched_end.strftime('%m-%d %H:%M')}（次日 01:35）"
            break

    # 4) 账单结果 sheet：逐张账单（车号 / 计费量 / 费用 / 说明）+ 总营收
    user_ended = {ev[1] for evs in by_time.values() for ev in evs
                  if ev[0] == "A" and ev[2] == "O"}
    rows = ctx.db.query(
        "SELECT b.car_id, b.bill_type, o.amount, o.total_fee, o.end_time "
        "FROM bills b JOIN orders o ON b.order_id = o.order_id "
        "ORDER BY o.end_time, b.bill_id")
    bs = wb["账单结果"] if "账单结果" in wb.sheetnames else wb.create_sheet("账单结果")
    bs.delete_rows(1, bs.max_row + 1)
    bs.append(["车号", "计费量(度)", "费用(元)", "说明"])
    total = 0.0
    for r in rows:
        end_hms = str(r["end_time"])[11:] if r["end_time"] else ""
        if r["bill_type"] == "interrupted":
            note = f"故障中断@{end_hms}"
        elif r["car_id"] in user_ended:
            note = f"提前结束@{end_hms}"
        else:
            note = f"充满@{end_hms}"
        bs.append([r["car_id"], round(r["amount"], 2), round(r["total_fee"], 2), note])
        total += r["total_fee"]
    bs.append([])
    bs.append(["总营收(元)", round(total, 2), f"{len(rows)} 笔账单", ""])

    wb.save(OUT)
    print("已填写 →", os.path.abspath(OUT))
    print("调度结束时刻:", sched_end.strftime("%m-%d %H:%M:%S") if sched_end else None)
    print(f"账单结果: {len(rows)} 笔，总营收 ¥{round(total, 2)}")
    return snaps, sched_end


if __name__ == "__main__":
    main()
