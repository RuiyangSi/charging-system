"""Replay the acceptance Excel events as system operations.

The workbook uses tuples shaped as:
    (event_type, target_id, charge_type, value)

This test keeps the acceptance data in one place: the Excel sheet remains the
source of truth, and the backend mapping is verified from that source.
"""
from __future__ import annotations

import re
import tempfile
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from app.context import AppContext
from app.domain.enums import BusinessError, PileStatus, RequestStatus

SIM_DAY = datetime(2026, 6, 11)
EVENT_RE = re.compile(r"^\(([ABC]),([^,]+),([FTO]),(-?\d+(?:\.\d+)?)\)$")


@dataclass(frozen=True)
class AcceptanceEvent:
    row: int
    at: datetime
    kind: str
    target: str
    mode: str
    value: float

    @property
    def raw(self) -> str:
        value = int(self.value) if self.value == int(self.value) else self.value
        return f"({self.kind},{self.target},{self.mode},{value})"

    @property
    def operation(self) -> str:
        if self.kind == "A" and self.mode in ("F", "T"):
            label = "快充" if self.mode == "F" else "慢充"
            return f"{self.target} 发起{label}请求，电量 {self.value:g} 度"
        if self.kind == "A" and self.mode == "O":
            return f"{self.target} 取消/结束充电"
        if self.kind == "B" and self.value == 0:
            return f"{self.target} 充电桩故障"
        if self.kind == "B" and self.value == 1:
            return f"{self.target} 充电桩恢复"
        if self.kind == "C":
            parts = []
            if self.mode in ("F", "T"):
                label = "快充" if self.mode == "F" else "慢充"
                parts.append(f"模式改为{label}")
            if self.value != -1:
                parts.append(f"电量改为 {self.value:g} 度")
            if not parts:
                parts.append("请求不变")
            return f"{self.target} 修改充电请求：" + "，".join(parts)
        return f"未知事件 {self.raw}"


def _workbook_path() -> Path:
    root = Path(__file__).resolve().parents[2]
    candidates = [p for p in root.glob("*.xlsx") if not p.name.startswith("~$")]
    assert candidates, "acceptance workbook not found"
    return candidates[0]


def _parse_events() -> list[AcceptanceEvent]:
    wb = openpyxl.load_workbook(_workbook_path(), data_only=True)
    ws = wb.worksheets[1]
    events: list[AcceptanceEvent] = []
    for row in range(1, ws.max_row + 1):
        time_value = ws.cell(row, 1).value
        event_value = ws.cell(row, 2).value
        if not time_value or not event_value:
            continue
        match = EVENT_RE.match(str(event_value).strip())
        if not match:
            continue
        hour, minute, second = [int(x) for x in str(time_value).strip().split(":")]
        events.append(
            AcceptanceEvent(
                row=row,
                at=SIM_DAY.replace(hour=hour, minute=minute, second=second),
                kind=match.group(1),
                target=match.group(2).strip(),
                mode=match.group(3),
                value=float(match.group(4)),
            )
        )
    return sorted(events, key=lambda e: (e.at, e.row))


def _apply_event(ctx: AppContext, event: AcceptanceEvent) -> None:
    if event.kind == "A" and event.mode in ("F", "T"):
        ctx.charging_service.submit_request(event.target, event.mode, event.value)
    elif event.kind == "A" and event.mode == "O":
        ctx.charging_service.cancel(event.target)
    elif event.kind == "B" and event.value == 0:
        ctx.schedule_service.handle_fault(event.target)
    elif event.kind == "B" and event.value == 1:
        pile = ctx.station.find_pile(event.target)
        if pile.status is PileStatus.FAULT:
            ctx.schedule_service.handle_recovery(event.target)
        else:
            ctx.pile_service.power_on(event.target)
            ctx.pile_service.run_pile(event.target)
    elif event.kind == "C":
        active = ctx.station.find_active_request(event.target)
        if event.mode in ("F", "T") and active and active.mode.value != event.mode:
            ctx.charging_service.modify_mode(event.target, event.mode)
        if event.value != -1:
            ctx.charging_service.modify_amount(event.target, event.value)


def _capture(ctx: AppContext) -> dict:
    now = ctx.clock.now()
    piles = {}
    for pile in ctx.station.charging_area.get_all_piles():
        cells = []
        for request in pile.queue.get_cars():
            if request.status is RequestStatus.CHARGING:
                estimate = ctx.billing_service.estimate(request, pile, now)
                cells.append(
                    f"({request.car_id},{estimate['chargedAmount']:.2f},"
                    f"{estimate['totalFee']:.2f})"
                )
            else:
                cells.append(f"({request.car_id},排队)")
        piles[pile.pile_id] = cells
    waiting = [
        f"({request.car_id},{request.mode.value},{request.requested_amount:.2f})"
        for request in ctx.station.waiting_area.get_all_requests()
    ]
    return {"piles": piles, "waiting": waiting}


def test_excel_events_are_parsed_and_mapped():
    events = _parse_events()
    assert len(events) == 42
    assert events[0].raw == "(A,V1,T,40)"
    assert events[-1].raw == "(B,F1,O,1)"
    assert events[0].operation == "V1 发起慢充请求，电量 40 度"
    assert [e.kind for e in events].count("B") == 4
    assert [e.kind for e in events].count("C") == 3


def test_replay_excel_events_with_waiting_area_capacity(ctx):
    ctx.config.data["interruptPolicy"] = "requeue"
    events = _parse_events()
    failures: list[tuple[AcceptanceEvent, str]] = []
    snapshots = {}

    for event in events:
        ctx.clock.set_time(event.at)
        ctx.step()
        try:
            _apply_event(ctx, event)
        except BusinessError as exc:
            failures.append((event, str(exc)))
        ctx.step()
        snapshots[event.at.strftime("%H:%M")] = _capture(ctx)

    # Existing sample cells in the workbook still match the simulation.
    assert snapshots["06:05"]["piles"]["T1"] == ["(V1,0.83,1.00)"]
    assert snapshots["06:05"]["piles"]["T2"] == ["(V2,0.00,0.00)"]
    assert snapshots["07:05"]["waiting"] == ["(V13,F,110.00)"]
    assert snapshots["07:10"]["waiting"] == ["(V13,F,110.00)", "(V14,F,95.00)"]
    assert snapshots["07:15"]["waiting"] == ["(V13,F,110.00)", "(V14,F,95.00)"]

    # With the corrected requirement N=10, the original workbook overbooks the
    # waiting area. These requests must be rejected cleanly with no active ghost
    # request left behind.
    capacity_failures = [
        event.target for event, message in failures if "等候区已满" in message
    ]
    missing_request_failures = [
        event.target for event, message in failures if "没有进行中的充电请求" in message
    ]
    assert capacity_failures == ["V26", "V27", "V28", "V30"]
    assert missing_request_failures == ["V27", "V28"]
    for target in set(capacity_failures):
        assert ctx.station.find_active_request(target) is None


def test_replay_excel_until_all_accepted_requests_finish(ctx):
    ctx = AppContext(tempfile.mkdtemp(prefix="excel_replay_"), db_path=":memory:")
    ctx.clock.set_speed(0)
    ctx.clock.set_time(SIM_DAY.replace(hour=6, minute=0, second=0))
    events = _parse_events()
    failures = []

    for event in events:
        ctx.clock.set_time(event.at)
        ctx.step()
        try:
            _apply_event(ctx, event)
        except BusinessError as exc:
            failures.append((event, str(exc)))
        ctx.step()

    current = events[-1].at
    deadline = SIM_DAY + timedelta(days=2)
    while current <= deadline:
        active = [r for r in ctx.station.requests.values() if r.status.active]
        if not active:
            break
        current += timedelta(minutes=1)
        ctx.clock.set_time(current)
        ctx.step()

    capacity_failures = [
        event.target for event, message in failures if "等候区已满" in message
    ]
    missing_request_failures = [
        event.target for event, message in failures if "没有进行中的充电请求" in message
    ]
    assert capacity_failures == ["V26", "V27", "V28", "V30"]
    assert missing_request_failures == ["V27", "V28"]
    assert not [r for r in ctx.station.requests.values() if r.status.active]
    assert ctx.billing_service.request_bill("V1")
