from __future__ import annotations

import csv
import json
import re
import sys
import tempfile
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from app.context import AppContext  # noqa: E402
from app.domain.enums import BusinessError, PileStatus, RequestStatus  # noqa: E402

SIM_DAY = datetime(2026, 6, 11)
EVENT_RE = re.compile(r"^\(([ABC]),([^,]+),([FTO]),(-?\d+(?:\.\d+)?)\)$")
PILE_IDS = ("F1", "F2", "T1", "T2", "T3")


def format_minutes_seconds(hours: float) -> str:
    total_seconds = max(0, round(hours * 3600))
    minutes, seconds = divmod(total_seconds, 60)
    return f"{minutes}分{seconds:02d}秒"


@dataclass(frozen=True)
class Event:
    index: int
    row: int
    time: str
    kind: str
    target: str
    mode: str
    value: float

    @property
    def at(self) -> datetime:
        hour, minute, second = [int(x) for x in self.time.split(":")]
        return SIM_DAY.replace(hour=hour, minute=minute, second=second)

    @property
    def raw(self) -> str:
        value = int(self.value) if self.value == int(self.value) else self.value
        return f"({self.kind},{self.target},{self.mode},{value})"

    @property
    def operation(self) -> str:
        if self.kind == "A" and self.mode in ("F", "T"):
            label = "fast" if self.mode == "F" else "trickle"
            return f"{self.target} submit {label} charge request, amount={self.value:g}"
        if self.kind == "A" and self.mode == "O":
            return f"{self.target} cancel/end charge"
        if self.kind == "B" and self.value == 0:
            return f"{self.target} pile fault"
        if self.kind == "B" and self.value == 1:
            return f"{self.target} pile recovery"
        if self.kind == "C":
            parts = []
            if self.mode in ("F", "T"):
                parts.append(f"mode={self.mode}")
            if self.value != -1:
                parts.append(f"amount={self.value:g}")
            return f"{self.target} modify request: " + ", ".join(parts)
        return f"unknown event {self.raw}"


def workbook_path() -> Path:
    exact = ROOT / "副本作业验收用例.xlsx"
    if exact.exists():
        return exact
    candidates = [
        p for p in ROOT.glob("*.xlsx")
        if not p.name.startswith("~$") and ".before_fill" not in p.name
    ]
    if not candidates:
        raise FileNotFoundError("No xlsx file found in project root")
    return candidates[0]


def parse_events() -> list[Event]:
    wb = openpyxl.load_workbook(workbook_path(), data_only=True)
    ws = wb.worksheets[1]
    events: list[Event] = []
    for row in range(1, ws.max_row + 1):
        time_value = ws.cell(row, 1).value
        event_value = ws.cell(row, 2).value
        if not time_value or not event_value:
            continue
        match = EVENT_RE.match(str(event_value).strip())
        if not match:
            continue
        events.append(
            Event(
                index=len(events) + 1,
                row=row,
                time=str(time_value).strip(),
                kind=match.group(1),
                target=match.group(2).strip(),
                mode=match.group(3),
                value=float(match.group(4)),
            )
        )
    return sorted(events, key=lambda e: (e.at, e.row))


def apply_event(ctx: AppContext, event: Event) -> None:
    if event.kind == "A" and event.mode in ("F", "T"):
        ctx.charging_service.submit_request(event.target, event.mode, event.value)
    elif event.kind == "A" and event.mode == "O":
        ctx.charging_service.cancel(event.target)
    elif event.kind == "B" and event.value == 0:
        ctx.schedule_service.handle_fault(event.target)
    elif event.kind == "B" and event.value == 1:
        pile = ctx.station.find_pile(event.target)
        if pile.status is PileStatus.FAULT:
            ctx.schedule_service.handle_recovery(event.target)
        else:
            ctx.pile_service.power_on(event.target)
            ctx.pile_service.run_pile(event.target)
    elif event.kind == "C":
        active = ctx.station.find_active_request(event.target)
        if event.mode in ("F", "T") and active and active.mode.value != event.mode:
            ctx.charging_service.modify_mode(event.target, event.mode)
        if event.value != -1:
            ctx.charging_service.modify_amount(event.target, event.value)


def request_row(ctx: AppContext, event: Event, location: str, request, pile=None, slot=None) -> dict:
    now = ctx.clock.now()
    power = pile.power if pile is not None else None
    charged = request.charged_amount(now, power) if power else request.actual_amount or 0.0
    fee = ""
    if pile is not None and request.status is RequestStatus.CHARGING:
        fee = ctx.billing_service.estimate(request, pile, now)["totalFee"]
    wait_time = max((now - request.submit_time).total_seconds() / 3600.0, 0.0)
    return {
        "event_index": event.index,
        "time": event.time,
        "event": event.raw,
        "location": location,
        "pile_id": pile.pile_id if pile is not None else "",
        "slot": slot if slot is not None else "",
        "car_id": request.car_id,
        "queue_number": request.queue_number,
        "mode": request.mode.value,
        "status": request.status.value,
        "requested_amount": round(request.requested_amount, 2),
        "charged_amount": round(charged, 2),
        "current_fee": round(fee, 2) if fee != "" else "",
        "submit_time": request.submit_time.strftime("%H:%M:%S"),
        "wait_time": format_minutes_seconds(wait_time),
    }


def capture(ctx: AppContext, event: Event, result: str, error: str) -> dict:
    now = ctx.clock.now()
    pile_rows = []
    pile_summary = {}
    for pile_id in PILE_IDS:
        pile = ctx.station.find_pile(pile_id)
        cells = []
        cars = pile.queue.get_cars()
        for idx in range(pile.queue.capacity):
            request = cars[idx] if idx < len(cars) else None
            slot_label = "charging" if idx == 0 else f"waiting_{idx}"
            if request is None:
                cells.append("-")
                pile_rows.append({
                    "event_index": event.index,
                    "time": event.time,
                    "event": event.raw,
                    "pile_id": pile_id,
                    "slot": idx + 1,
                    "slot_type": slot_label,
                    "car_id": "",
                    "queue_number": "",
                    "mode": "",
                    "status": "EMPTY",
                    "requested_amount": "",
                    "charged_amount": "",
                    "current_fee": "",
                    "wait_time": "",
                })
                continue
            row = request_row(ctx, event, "pile", request, pile=pile, slot=idx + 1)
            row["slot_type"] = slot_label
            pile_rows.append(row)
            if request.status is RequestStatus.CHARGING:
                fee = ctx.billing_service.estimate(request, pile, now)["totalFee"]
                cells.append(f"({request.car_id},{request.charged_amount(now, pile.power):.2f},{fee:.2f})")
            else:
                cells.append(f"({request.car_id},queue,{request.requested_amount:.2f})")
        pile_summary[pile_id] = " | ".join(cells)

    waiting_rows = [
        request_row(ctx, event, "waiting_area", request, slot=idx + 1)
        for idx, request in enumerate(ctx.station.waiting_area.get_all_requests())
    ]
    priority_rows = [
        request_row(ctx, event, "priority_waiting", request, slot=idx + 1)
        for idx, request in enumerate(getattr(ctx.schedule_service, "priority_waiting", []))
    ]
    waiting_summary = " | ".join(
        f"({r['car_id']},{r['mode']},{r['requested_amount']:.2f})" for r in waiting_rows
    ) or "-"
    priority_summary = " | ".join(
        f"({r['car_id']},{r['mode']},{r['requested_amount']:.2f})" for r in priority_rows
    ) or "-"
    return {
        "operation": {
            **asdict(event),
            "event": event.raw,
            "operation": event.operation,
            "result": result,
            "error": error,
        },
        "snapshot": {
            "event_index": event.index,
            "time": event.time,
            "event": event.raw,
            "result": result,
            "error": error,
            **pile_summary,
            "waiting_area": waiting_summary,
            "priority_waiting": priority_summary,
            "waiting_count": len(waiting_rows),
            "priority_waiting_count": len(priority_rows),
        },
        "pile_rows": pile_rows,
        "waiting_rows": waiting_rows,
        "priority_rows": priority_rows,
    }


def replay() -> dict:
    events = parse_events()
    ctx = AppContext(tempfile.mkdtemp(prefix="excel_replay_export_"), db_path=":memory:")
    ctx.clock.set_speed(0)
    ctx.clock.set_time(SIM_DAY.replace(hour=6, minute=0, second=0))
    ctx.config.data["interruptPolicy"] = "requeue"

    operations = []
    snapshots = []
    pile_rows = []
    waiting_rows = []
    priority_rows = []
    failures = []

    for event in events:
        ctx.clock.set_time(event.at)
        ctx.step()
        result = "OK"
        error = ""
        try:
            apply_event(ctx, event)
        except BusinessError as exc:
            result = "FAILED"
            error = str(exc)
            failures.append({"time": event.time, "event": event.raw, "error": error})
        ctx.step()
        captured = capture(ctx, event, result, error)
        operations.append(captured["operation"])
        snapshots.append(captured["snapshot"])
        pile_rows.extend(captured["pile_rows"])
        waiting_rows.extend(captured["waiting_rows"])
        priority_rows.extend(captured["priority_rows"])

    current = events[-1].at
    deadline = SIM_DAY + timedelta(days=2)
    while current <= deadline:
        active = [request for request in ctx.station.requests.values() if request.status.active]
        if not active:
            break
        current += timedelta(minutes=1)
        ctx.clock.set_time(current)
        ctx.step()

    return {
        "source": str(workbook_path()),
        "event_count": len(events),
        "failure_count": len(failures),
        "failures": failures,
        "schedule_end": current.strftime("%Y-%m-%d %H:%M:%S"),
        "operations": operations,
        "snapshots": snapshots,
        "pile_vehicles": pile_rows,
        "waiting_vehicles": waiting_rows,
        "priority_waiting_vehicles": priority_rows,
    }


def write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    headers = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(key)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, data: dict) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheets = {
        "Operations": data["operations"],
        "Snapshots": data["snapshots"],
        "PileVehicles": data["pile_vehicles"],
        "WaitingVehicles": data["waiting_vehicles"],
        "PriorityWaiting": data["priority_waiting_vehicles"] or [
            {"event_index": "", "time": "", "event": "", "car_id": ""}
        ],
        "Failures": data["failures"] or [{"time": "", "event": "", "error": ""}],
    }
    for title, rows in sheets.items():
        ws = wb.create_sheet(title)
        headers = list(rows[0].keys()) if rows else []
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        ws.freeze_panes = "A2"
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 50)
    wb.save(path)


def main() -> None:
    output = ROOT / "output"
    output.mkdir(exist_ok=True)
    data = replay()
    json_path = output / "excel_replay_results.json"
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(output / "excel_replay_operations.csv", data["operations"])
    write_csv(output / "excel_replay_snapshots.csv", data["snapshots"])
    write_csv(output / "excel_replay_pile_vehicles.csv", data["pile_vehicles"])
    write_csv(output / "excel_replay_waiting_vehicles.csv", data["waiting_vehicles"])
    write_csv(output / "excel_replay_priority_waiting_vehicles.csv", data["priority_waiting_vehicles"])
    write_xlsx(output / "excel_replay_results.xlsx", data)

    print(f"source={data['source']}")
    print(f"events={data['event_count']} failures={data['failure_count']}")
    print(f"schedule_end={data['schedule_end']}")
    print(f"wrote={output / 'excel_replay_results.xlsx'}")
    print(f"wrote={json_path}")


if __name__ == "__main__":
    main()
