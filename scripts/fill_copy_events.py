import json
import openpyxl
from datetime import time

src_json = 'output/event_summary.json'
src_xlsx = '副本作业验收用例.xlsx'
out_xlsx = '副本作业验收用例_filled.xlsx'

with open(src_json, 'r', encoding='utf-8') as f:
    timeline = json.load(f)

wb = openpyxl.load_workbook(src_xlsx)
ws = wb['测试用例']

# build map time_str -> entry
# keep timeline list and provide flexible lookup by time string
timeline_list = timeline


def find_entry_for_time(tstr):
    if not tstr:
        return None
    # exact match
    for e in timeline_list:
        if e.get('time') == tstr:
            return e
    # try startswith (timeline may have seconds)
    for e in timeline_list:
        if e.get('time') and e.get('time').startswith(tstr):
            return e
    # try matching by HH:MM prefix
    short = tstr[:5]
    for e in timeline_list:
        if e.get('time') and e.get('time')[:5] == short:
            return e
    return None

# column index mapping based on template (0-based):
# 0: 时刻, 1: 事件, 2: 快充1, 3: 快充2, 4: 慢充1, 5: 慢充2, 6: 慢充3, 7..: 等候区
fast_idxs = [2, 3]
slow_idxs = [4, 5, 6]
wait_start = 7
max_wait = 10

# iterate rows starting at row 3 (as seen)
def set_cell_safe(ws, row_idx, col_idx, val):
    try:
        ws.cell(row=row_idx, column=col_idx).value = val
    except AttributeError:
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row_idx <= merged.max_row and merged.min_col <= col_idx <= merged.max_col:
                ws.cell(row=merged.min_row, column=merged.min_col).value = val
                return
        raise


for row_idx in range(3, ws.max_row + 1):
    time_cell = ws.cell(row=row_idx, column=1).value
    if time_cell is None:
        continue
    if isinstance(time_cell, time):
        tstr = time_cell.strftime('%H:%M:%S')
    else:
        tstr = str(time_cell)
    entry = find_entry_for_time(tstr)
    if not entry:
        continue

    # write event_raw into column 2
    if entry.get('event_raw'):
        set_cell_safe(ws, row_idx, 1, entry['event_raw'])

    # fill fast columns
    fast_list = entry.get('fast') or []
    for i, idx in enumerate(fast_idxs):
        val = fast_list[i] if i < len(fast_list) else None
        set_cell_safe(ws, row_idx, idx, val)

    # fill slow columns
    slow_list = entry.get('slow') or []
    for i, idx in enumerate(slow_idxs):
        val = slow_list[i] if i < len(slow_list) else None
        set_cell_safe(ws, row_idx, idx, val)

    # fill waiting columns
    waiting_list = entry.get('waiting') or []
    for j in range(max_wait):
        col_idx = wait_start + j
        val = waiting_list[j] if j < len(waiting_list) else None
        set_cell_safe(ws, row_idx, col_idx, val)

wb.save(out_xlsx)
print('Wrote filled workbook to', out_xlsx)
