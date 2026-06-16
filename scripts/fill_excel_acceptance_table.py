from __future__ import annotations

import re
import shutil
import sys
import tempfile
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from app.context import AppContext  # noqa: E402
from app.domain.enums import BusinessError, PileStatus, RequestStatus  # noqa: E402

SIM_DAY = datetime(2026, 6, 11)
EVENT_RE = re.compile(r"^\(([ABC]),([^,]+),([FTO]),(-?\d+(?:\.\d+)?)\)$")
PILE_COLS = {"F1": 3, "F2": 4, "T1": 5, "T2": 6, "T3": 7}
WAIT_COL = 8
NOTE_COL = 9


def workbook_path() -> Path:
    exact = ROOT / "副本作业验收用例.xlsx"
    if exact.exists():
        return exact
    candidates = [
        p for p in ROOT.glob("*.xlsx")
        if not p.name.startswith("~$") and ".before_fill" not in p.name
    ]
    if not candidates:
        raise FileNotFoundError("No xlsx file found in project root")
    # Prefer the original acceptance copy over generated outputs.
    return sorted(candidates, key=lambda p: ("副本" not in p.name, p.name))[0]


def parse_event(value):
    match = EVENT_RE.match(str(value or "").strip())
    if not match:
        return None
    raw_value = float(match.group(4))
    return match.group(1), match.group(2).strip(), match.group(3), raw_value


def parse_time(value):
    text = str(value or "").strip()
    match = re.match(r"^(\d{1,2}):(\d{2}):(\d{2})$", text)
    if not match:
        return None
    hour, minute, second = [int(x) for x in match.groups()]
    return SIM_DAY.replace(hour=hour, minute=minute, second=second)


def apply_event(ctx: AppContext, event):
    kind, target, mode, value = event
    if kind == "A" and mode in ("F", "T"):
        ctx.charging_service.submit_request(target, mode, value)
    elif kind == "A" and mode == "O":
        ctx.charging_service.cancel(target)
    elif kind == "B" and value == 0:
        ctx.schedule_service.handle_fault(target)
    elif kind == "B" and value == 1:
        pile = ctx.station.find_pile(target)
        if pile.status is PileStatus.FAULT:
            ctx.schedule_service.handle_recovery(target)
        else:
            ctx.pile_service.power_on(target)
            ctx.pile_service.run_pile(target)
    elif kind == "C":
        active = ctx.station.find_active_request(target)
        if mode in ("F", "T") and active and active.mode.value != mode:
            ctx.charging_service.modify_mode(target, mode)
        if value != -1:
            ctx.charging_service.modify_amount(target, value)


def render_pile(ctx: AppContext, pile_id: str, now: datetime) -> list[str]:
    pile = ctx.station.find_pile(pile_id)
    if pile.status is PileStatus.FAULT:
        return ["故障", "故障", "故障"]
    cells: list[str] = []
    for request in pile.queue.get_cars():
        if request.status is RequestStatus.CHARGING:
            estimate = ctx.billing_service.estimate(request, pile, now)
            cells.append(
                f"({request.car_id},{estimate['chargedAmount']:.2f},"
                f"{estimate['totalFee']:.2f})"
            )
        else:
            # Template convention: queued cars behind the charging slot show
            # zero charged amount and zero current fee.
            cells.append(f"({request.car_id},0.00,0.00)")
    while len(cells) < pile.queue.capacity:
        cells.append("-")
    return cells[:pile.queue.capacity]


def render_waiting(ctx: AppContext) -> str:
    waiting = ctx.station.waiting_area.get_all_requests()
    if not waiting:
        return ""
    return "-".join(
        f"({request.car_id},{request.mode.value},{request.requested_amount:.2f})"
        for request in waiting
    )


def render_priority_waiting(ctx: AppContext) -> str:
    cars = getattr(ctx.schedule_service, "priority_waiting", [])
    if not cars:
        return ""
    return "故障重调度队列：" + "-".join(
        f"({request.car_id},{request.mode.value},{request.requested_amount:.2f})"
        for request in cars
    )


def main() -> None:
    path = workbook_path()
    backup = path.with_suffix(".before_fill.xlsx")
    if not backup.exists():
        shutil.copy2(path, backup)

    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[1]

    event_rows = []
    for row in range(1, ws.max_row + 1):
        at = parse_time(ws.cell(row, 1).value)
        event = parse_event(ws.cell(row, 2).value)
        if at and event:
            event_rows.append((row, at, event))
    if not event_rows:
        raise RuntimeError("No acceptance events found")

    ctx = AppContext(tempfile.mkdtemp(prefix="excel_fill_"), db_path=":memory:")
    ctx.clock.set_speed(0)
    ctx.clock.set_time(SIM_DAY.replace(hour=6, minute=0, second=0))
    ctx.config.data["interruptPolicy"] = "requeue"

    failures = []
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    for row, at, event in sorted(event_rows, key=lambda item: (item[1], item[0])):
        ctx.clock.set_time(at)
        ctx.step()
        note = ""
        try:
            apply_event(ctx, event)
        except BusinessError as exc:
            note = f"执行失败：{exc}"
            failures.append((row, at.strftime("%H:%M:%S"), event, str(exc)))
        ctx.step()
        now = ctx.clock.now()

        for pile_id, col in PILE_COLS.items():
            for offset, text in enumerate(render_pile(ctx, pile_id, now)):
                cell = ws.cell(row + offset, col)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = text
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        wait_cell = ws.cell(row, WAIT_COL)
        wait_cell.value = render_waiting(ctx)
        wait_cell.alignment = Alignment(vertical="center", wrap_text=True)

        note_cell = ws.cell(row, NOTE_COL)
        priority_note = render_priority_waiting(ctx)
        note_cell.value = "；".join(part for part in (note, priority_note) if part)
        if note_cell.value:
            note_cell.fill = note_fill
            note_cell.alignment = Alignment(vertical="center", wrap_text=True)

    current = max(at for _, at, _ in event_rows)
    deadline = SIM_DAY + timedelta(days=2)
    while current <= deadline:
        active = [request for request in ctx.station.requests.values() if request.status.active]
        if not active:
            break
        current += timedelta(minutes=1)
        ctx.clock.set_time(current)
        ctx.step()

    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 2).value or "").strip() == "调度结束":
            days = (current.date() - SIM_DAY.date()).days
            ws.cell(row, 1).value = f"{current.strftime('%H:%M')}+{days}"
            for pile_id, col in PILE_COLS.items():
                for offset in range(3):
                    cell = ws.cell(row + offset, col)
                    if not isinstance(cell, MergedCell):
                        cell.value = "-"
            ws.cell(row, WAIT_COL).value = ""
            ws.cell(row, NOTE_COL).value = f"实际调度结束：{current.strftime('%Y-%m-%d %H:%M:%S')}"
            break

    ws.cell(2, NOTE_COL).value = "执行备注"
    ws.cell(2, NOTE_COL).alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["H"].width = 55
    ws.column_dimensions["I"].width = 38

    try:
        wb.save(path)
        filled_path = path
    except PermissionError:
        filled_path = path.with_name(f"{path.stem}_priority_filled{path.suffix}")
        wb.save(filled_path)
    print(f"filled={filled_path}")
    print(f"backup={backup}")
    print(f"events={len(event_rows)} failures={len(failures)}")
    print(f"schedule_end={current.strftime('%Y-%m-%d %H:%M:%S')}")
    for row, at, event, message in failures:
        value = int(event[3]) if event[3] == int(event[3]) else event[3]
        print(f"failure row={row} time={at} event=({event[0]},{event[1]},{event[2]},{value}) message={message}")


if __name__ == "__main__":
    main()
