import openpyxl

path = '副本作业验收用例.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
print('SHEETS:', wb.sheetnames)
for name in wb.sheetnames:
    print('\n--- Sheet:', name, '---')
    ws = wb[name]
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(i, row)
        if i >= 40:
            break
