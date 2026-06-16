import sys
try:
    import openpyxl
except Exception as e:
    print('MISSING_OPENPYXL')
    sys.exit(0)

path = 'docs/作业验收用例_本组填写.xlsx'
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print('SHEETS:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('\n--- Sheet:', name, '---')
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(i, row)
        if i >= 20:
            break
