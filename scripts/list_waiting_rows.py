import openpyxl
wb=openpyxl.load_workbook('副本作业验收用例_filled.xlsx', data_only=True)
ws=wb['测试用例']
res=[]
for r in range(3, ws.max_row+1):
    t=ws.cell(row=r,column=1).value
    w=ws.cell(row=r,column=8).value
    if w:
        res.append((r,t,w))
for item in res[:50]:
    print(item)
print('total waiting rows:', len(res))
