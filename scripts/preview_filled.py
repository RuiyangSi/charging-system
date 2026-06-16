import openpyxl

path = '副本作业验收用例_filled_ids.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['测试用例']

max_row = min(ws.max_row, 40)
max_col = min(ws.max_column, 20)

for r in range(1, max_row+1):
    vals = [ws.cell(row=r, column=c).value for c in range(1, max_col+1)]
    print(r, vals)
