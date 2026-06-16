import openpyxl
wb=openpyxl.load_workbook('副本作业验收用例_filled.xlsx', data_only=True)
ws=wb['测试用例']
for r in range(1, ws.max_row+1):
    for c in range(1, ws.max_column+1):
        v=ws.cell(row=r,column=c).value
        if v and 'V13' in str(v):
            print('row',r,'col',c,'val',v)
