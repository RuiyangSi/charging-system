import re
import openpyxl

src = '副本作业验收用例_filled.xlsx'
out = '副本作业验收用例_filled_ids.xlsx'

wb = openpyxl.load_workbook(src)
ws = wb['测试用例']

id_re = re.compile(r"(V\d+|F\d+|T\d+)")

# column indices (0-based) as used: 0 time,1 event,2 fast1,3 fast2,4 slow1,5 slow2,6 slow3,7.. waiting
fast_idxs = [2, 3]
slow_idxs = [4, 5, 6]
wait_start = 7
max_wait = 10

for row_idx in range(3, ws.max_row + 1):
    # time
    time_cell = ws.cell(row=row_idx, column=1).value
    if time_cell is None:
        continue

    # process fast
    for col in fast_idxs:
        cell = ws.cell(row=row_idx, column=col + 1)
        val = cell.value
        if not val:
            cell.value = None
            continue
        m = id_re.search(str(val))
        cell.value = m.group(1) if m else None

    # process slow
    for col in slow_idxs:
        cell = ws.cell(row=row_idx, column=col + 1)
        val = cell.value
        if not val:
            cell.value = None
            continue
        m = id_re.search(str(val))
        cell.value = m.group(1) if m else None

    # process waiting
    for j in range(max_wait):
        col = wait_start + j
        if col + 1 > ws.max_column:
            break
        cell = ws.cell(row=row_idx, column=col + 1)
        val = cell.value
        if not val:
            cell.value = None
            continue
        m = id_re.search(str(val))
        cell.value = m.group(1) if m else None

wb.save(out)
print('Wrote IDs-only workbook to', out)
