import openpyxl
import re
import csv
import json
from pathlib import Path


def parse_tuple_cell(cell_val):
    if not cell_val:
        return None
    s = str(cell_val).strip()
    if s in {'-', ''}:
        return None
    m = re.match(r"^\(([^)]+)\)", s)
    if not m:
        # try to find a vehicle id inside
        mv = re.search(r"(V\d+|F\d+|T\d+)", s)
        return (None, mv.group(1), None, None) if mv else None
    parts = [p.strip() for p in m.group(1).split(',')]
    if len(parts) != 4:
        return None
    op, vid, mode, val = parts
    try:
        if '.' in val:
            valn = float(val)
        else:
            valn = int(val)
    except Exception:
        valn = val
    return (op, vid, mode, valn)


def summarize(path='docs/作业验收用例_本组填写.xlsx', out_csv='output/event_summary.csv'):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['测试用例']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]

    fast_cols = [i for i, h in enumerate(header) if h and isinstance(h, str) and '快充' in h]
    slow_cols = [i for i, h in enumerate(header) if h and isinstance(h, str) and h.startswith('慢充')]
    wait_start = next((i for i, h in enumerate(header) if h and isinstance(h, str) and '等候区' in h), None)

    timeline = []
    for r in rows[2:]:
        if not r or r[0] is None:
            continue
        t = r[0]
        event_raw = r[1]
        event = None
        if isinstance(event_raw, str) and event_raw.strip().startswith('('):
            event = parse_tuple_cell(event_raw)

        # collect fast
        fast_list = []
        for i in fast_cols:
            if i < len(r):
                parsed = parse_tuple_cell(r[i])
                if parsed:
                    fast_list.append(parsed[1])
                elif r[i] and str(r[i]).strip() not in {'-', ''}:
                    fast_list.append(str(r[i]).strip())
        slow_list = []
        for i in slow_cols:
            if i < len(r):
                parsed = parse_tuple_cell(r[i])
                if parsed:
                    slow_list.append(parsed[1])
                elif r[i] and str(r[i]).strip() not in {'-', ''}:
                    slow_list.append(str(r[i]).strip())
        waiting_list = []
        if wait_start is not None:
            for i in range(wait_start, len(r)):
                cell = r[i]
                parsed = parse_tuple_cell(cell)
                if parsed:
                    waiting_list.append(parsed[1])
                elif cell and str(cell).strip() not in {'-', ''}:
                    # try extract Vxx
                    m = re.search(r"(V\d+|F\d+|T\d+)", str(cell))
                    waiting_list.append(m.group(1) if m else str(cell).strip())

        entry = {
            'time': str(t),
            'event_raw': str(event_raw) if event_raw is not None else None,
            'event': {'op': event[0], 'id': event[1], 'mode': event[2], 'val': event[3]} if event else None,
            'fast': fast_list,
            'slow': slow_list,
            'waiting': waiting_list,
        }
        timeline.append(entry)

        # print succinct summary
        print('Time:', t)
        print('  Event:', event_raw)
        print('  Fast charging:', fast_list)
        print('  Slow charging:', slow_list)
        print('  Waiting:', waiting_list)
        print('')

    # write CSV and JSON
    Path('output').mkdir(exist_ok=True)
    with open(out_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['time', 'event_raw', 'fast_charging', 'slow_charging', 'waiting'])
        for e in timeline:
            writer.writerow([e['time'], e['event_raw'] or '', ';'.join(e['fast']), ';'.join(e['slow']), ';'.join(e['waiting'])])

    json_path = 'output/event_summary.json'
    with open(json_path, 'w', encoding='utf-8') as jf:
        json.dump(timeline, jf, ensure_ascii=False, indent=2)
    print('Wrote JSON to', json_path)
    return timeline


def get_state_at(query_time, timeline=None, path='output/event_summary.json'):
    # query_time should be string matching 'HH:MM:SS' or exact time string
    if timeline is None:
        import json as _json
        with open(path, 'r', encoding='utf-8') as f:
            timeline = _json.load(f)
    for e in timeline:
        if e['time'] == query_time:
            return e
    # try partial match (e.g., '06:20')
    for e in timeline:
        if e['time'].startswith(query_time):
            return e
    return None


def write_xlsx(timeline, out_xlsx='output/event_summary.xlsx'):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'EventSummary'
    headers = ['time', 'event_raw', 'event_op', 'event_id', 'event_mode', 'event_val', 'fast', 'slow', 'waiting']
    ws.append(headers)
    for e in timeline:
        ev = e.get('event') or {}
        row = [
            e.get('time'),
            e.get('event_raw'),
            ev.get('op') if ev else None,
            ev.get('id') if ev else None,
            ev.get('mode') if ev else None,
            ev.get('val') if ev else None,
            ';'.join(e.get('fast') or []),
            ';'.join(e.get('slow') or []),
            ';'.join(e.get('waiting') or []),
        ]
        ws.append(row)
    wb.save(out_xlsx)
    print('Wrote XLSX to', out_xlsx)
    return out_xlsx


if __name__ == '__main__':
    import argparse

    p = argparse.ArgumentParser()
    p.add_argument('--json', action='store_true', help='Write JSON and CSV output')
    p.add_argument('--query', type=str, help='Query state at given time string')
    args = p.parse_args()

    tl = summarize()
    # always write XLSX for convenience
    write_xlsx(tl)
    if args.query:
        st = get_state_at(args.query, timeline=tl)
        print('\nQuery result for', args.query, ':')
        print(st)
